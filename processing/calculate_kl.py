import json
import os
import sys
import csv
import traceback
import matplotlib.pyplot as plt
import numpy
import pandas as pd
import xlwt
from openpyxl import Workbook, load_workbook
import copy
import operator
import ntpath
import scipy
import numpy as np
from time import gmtime, strftime

from sklearn.feature_extraction.text import CountVectorizer

from utils.logger import Logger
import lexrank_config

TOP_K_WORDS = lexrank_config.calculate_kl['TOP_K_WORDS']            # present top words (30)
SMOOTHING_FACTOR = lexrank_config.calculate_kl['SMOOTHING_FACTOR']  # smoothing factor for calculate term contribution (1.0)
NGRAM_RANGE = lexrank_config.calculate_kl['NGRAM_RANGE']
VOCABULARY_METHOD = lexrank_config.calculate_kl['VOCABULARY_METHOD']  # KL methods: aggregate all documents

CONTRIBUTE_TYPE = lexrank_config.calculate_kl['CONTRIBUTE_TYPE']
NORMALIZE_CONTRIBUTE_FLAG = lexrank_config.calculate_kl['NORMALIZE_CONTRIBUTE']['flag']
NORMALIZE_CONTRIBUTE_TYPE = lexrank_config.calculate_kl['NORMALIZE_CONTRIBUTE']['type']         # TODO not in use

CLIP_DESCRIPTION = lexrank_config.calculate_kl['CLIP_DESCRIPTION']

FIND_WORD_DESCRIPTION_FLAG = lexrank_config.calculate_kl['FIND_WORD_DESCRIPTION']['flag']
FIND_WORD_DESCRIPTION_K = lexrank_config.calculate_kl['FIND_WORD_DESCRIPTION']['k']             # TODO not in use

KL_TYPE = lexrank_config.calculate_kl['KL_TYPE']['type']                    # KL for pos or words
KL_TARGET_COLUMN = lexrank_config.calculate_kl['KL_TYPE']['column_name']    # target column name contains the descriptions

PERSONALITY_TRAIT = lexrank_config.personality_trait        # list of personality traits

KL_CONFIGURATION_DICT = lexrank_config.calculate_kl
RESULT_DIR = '../results/data/kl/'


class CalculateKL:

    """
        calculate KL-contribute between to verticals/traits distribution (e.g. vocabulary of high/low extroversion)
        save two excel files with the most influence word (KL "contribution") regards to each h/l.

    Args:
        description_file_p: first group - csv with description
        description_file_q: second group - csv with description
        trait:
        vertical:
        p_title=None
        q_title=None
        ngram_range=(1,2): token n-gram to check contribute (first build distribution using them, than check importance)

    Returns:
        1. ../results/kl//all_words_contribute/cur_trait/____
            excel file with all words contribute between two distribution (e.g. extroversion/introversion)

        2. ../results/kl/all_words_contribute_detailed/cur_trait/____
            excel file with top k words contribute with additional data between two distribution

        3. ../results/kl/token/cur_trait/____
            excel file, one file to each token with all his description he appears in

    Raises:

    """

    def __init__(self, merge_df_path, log_file_name, level='info'):

        Logger.set_handlers('CalculateKL', log_file_name, level=level)

        CalculateKL.check_input()

        self.merge_df_path = merge_df_path

        self.file_name_p = None
        self.file_name_q = None

        self.cur_time = strftime("%Y-%m-%d %H:%M:%S", gmtime())         # global time to all folders

        # token and all the description he appears in
        self.dir_excel_token_appearance = '{}token/{}/'.format(RESULT_DIR, self.cur_time)

        # input to Lex-rank algorithm (word + value)
        self.dir_all_words_contribute = '{}all_words_contribute/{}/'.format(RESULT_DIR, self.cur_time)

        # top k words with highest contribute and further explanation
        self.dir_kl_detailed_word = '{}kl_detailed/{}/'.format(RESULT_DIR, self.cur_time)

        self.pt = None

        self.corpus = list()        # corpus contain two languagas models
        self.vectorizer = list()    # transform words into vectors of numbers using sklearn
        self.X = list()             # sparse vector represent vocabulary words
        self.X_dense = list()       # dense vector represent vocabulary words

        self.X_dense_p = list()     # dense tf for p distribution
        self.X_dense_q = list()     # dense tf for q distribution
        self.X_dense_kl = list()    # matrix contain two dense vectors for kl

        self.X_dense_binary_p = np.matrix([])
        self.X_dense_binary_q = np.matrix([])

        self.text_list_list_p = list()
        self.text_list_list_q = list()

        # documents method
        self.count_vec_p = CountVectorizer()
        self.count_vec_q = CountVectorizer()
        self.occurrence_doc_sum_p = list()
        self.occurrence_doc_sum_q = list()
        self.len_p = int                        # number of posts
        self.len_q = int                        # number of posts
        self.unique_buyer_p = int
        self.unique_buyer_q = int

        self.normalize_q = list()   # normalize
        self.normalize_p = list()
        csv.field_size_limit(sys.maxsize)

    def init_debug_log(self):

        file_prefix = 'calculate_kl'
        log_file_name = '../log/{}_{}.log'.format(file_prefix, self.cur_time)
        Logger.set_handlers('CalculateKL', log_file_name, level='debug')

    @staticmethod
    def check_input():
        if VOCABULARY_METHOD not in ['documents', 'aggregation']:
            raise ValueError('vocabulary method: {} is not defined'.format(str(VOCABULARY_METHOD)))

        if NGRAM_RANGE[0] > NGRAM_RANGE[1]:
            raise ValueError('n gram is not valid: {}'.format(str(NGRAM_RANGE)))

        if KL_TYPE not in ['words', 'POS']:
            raise ValueError('KL type is not defined: {}'.format(str(KL_TYPE)))

        if CONTRIBUTE_TYPE not in ['description_fraction', 'buyers_fraction']:
            raise ValueError('Contribute calculation is not defined: {}'.format(str(CONTRIBUTE_TYPE)))

        if (KL_TYPE == 'words' and KL_TARGET_COLUMN == 'description_POS_str') or \
                (KL_TYPE == 'POS' and KL_TARGET_COLUMN != 'description_POS_str'):
            raise ValueError('KL column name is not match to KL type defined: {} {}'.format
                             (str(KL_TYPE), str(KL_TARGET_COLUMN)))

        if not os.path.exists(RESULT_DIR):
            os.makedirs(RESULT_DIR)

    # main function - load vocabularies regards to vocabulary method
    def run_kl(self):

        if VOCABULARY_METHOD == 'documents':

            # iterate over all personality traits and calculate KL
            for pt in PERSONALITY_TRAIT:
                self.pt = pt
                self._init_class_variables(pt)                  # init all class variables due to 'for' loop
                self._load_vocabulary_groups(pt)                # extract two df regards to column value
                self.load_vocabulary_vector_documents()         # load documents + save attributes to later computation
                self.calculate_kl_and_language_models_documents()

            self._save_config_values()  # dump configuration

        elif VOCABULARY_METHOD == 'aggregation':
            raise ValueError('currently only vocabulary method documents is supported')
            # self.load_vocabulary_vector_aggregation()
            # self.calculate_kl_and_language_models_aggregation()

    def _init_class_variables(self, pt):

        # token and all the description he appears in
        self.dir_excel_token_appearance = '{}{}/token/'.format(RESULT_DIR, self.cur_time)

        # input to Lex-rank algorithm (word + value)
        self.dir_all_words_contribute = '{}{}/all_words_contribute/'.format(RESULT_DIR, self.cur_time)

        # top k words with highest contribute and further explanation
        self.dir_kl_detailed_word = '{}{}/kl_detailed/'.format(RESULT_DIR, self.cur_time)

        self.dir_config = '{}{}/'.format(RESULT_DIR, self.cur_time)

        self.corpus = list()  # corpus contain two languagas models
        self.vectorizer = list()  # transform words into vectors of numbers using sklearn
        self.X = list()  # sparse vector represent vocabulary words
        self.X_dense = list()  # dense vector represent vocabulary words

        self.X_dense_p = list()  # dense tf for p distribution
        self.X_dense_q = list()  # dense tf for q distribution
        self.X_dense_kl = list()  # matrix contain two dense vectors for kl

        self.X_dense_binary_p = np.matrix([])
        self.X_dense_binary_q = np.matrix([])

        self.text_list_list_p = list()
        self.text_list_list_q = list()

        # documents method
        self.count_vec_p = CountVectorizer()
        self.count_vec_q = CountVectorizer()
        self.occurrence_doc_sum_p = list()
        self.occurrence_doc_sum_q = list()
        self.len_p = int  # number of posts
        self.len_q = int  # number of posts
        self.unique_buyer_p = int
        self.unique_buyer_q = int
        self.normalize_q = list()  # normalize
        self.normalize_p = list()

        self.df_p = pd.DataFrame
        self.df_q = pd.DataFrame

    def _load_vocabulary_groups(self, pt):
        """
        split data regards to specific personality trait to H and L values
        :argument: pt personality trait string name
        :return: two DF - one to each group
        """

        merge_df = pd.read_csv(self.merge_df_path)
        # merge_df = merge_df.head(500)                 # for debugging mode

        dict_personality_values_df = {k: v for k, v in merge_df.groupby('{}_group'.format(pt))}

        for group_type, group in dict_personality_values_df.items():

            # log statistics and save histogram
            item_amount_per_users = group['buyer_id'].value_counts().tolist()

            user_amount = group['buyer_id'].nunique()       # number of unique users in the group (H/L/M)
            desc_num = group.shape[0]                       # number of description in the group

            self._plot_histogram(np.array(item_amount_per_users), user_amount, desc_num, group_type, pt=pt)

            Logger.info('{}: unique users: {}, num description: {}'.format(
                group_type,
                user_amount,
                desc_num))

        self.text_list_list_p = dict_personality_values_df['H'][KL_TARGET_COLUMN]      # .tolist()
        self.df_p = dict_personality_values_df['H']
        self.file_name_p = 'High'

        self.text_list_list_q = dict_personality_values_df['L'][KL_TARGET_COLUMN]      # .tolist()
        self.df_q = dict_personality_values_df['L']
        self.file_name_q = 'Low'

    def load_vocabulary_vector_documents(self):
        """ load vocabulary support aggregation method - KLPost """
        self.len_p = np.float(len(self.text_list_list_p))
        self.unique_buyer_p = np.float(len(self.df_p['buyer_id'].value_counts().index))
        Logger.info('P #of items descriptions: {}, unique buyers: {}'.format(str(self.len_p), self.unique_buyer_p))

        self.len_q = np.float(len(self.text_list_list_q))
        self.unique_buyer_q = np.float(len(self.df_q['buyer_id'].value_counts().index))
        Logger.info('Q #of items descriptions: {}, unique buyers: {}'.format(str(self.len_q), self.unique_buyer_q))

        #  P distribution
        self.count_vec_p = CountVectorizer(
            ngram_range=NGRAM_RANGE,
            stop_words='english',
            lowercase=False
        )
        X_train_counts_p = self.count_vec_p.fit_transform(self.text_list_list_p)   # count occurrence per word in documents
        X_dense_p = scipy.sparse.csr_matrix.todense(X_train_counts_p)   # dense transformation
        X_dense_binary_p = scipy.sign(X_dense_p)                        # binary - count 1-0 occurrence per documents
        self.occurrence_doc_sum_p = numpy.sum(X_dense_binary_p, axis=0).transpose()     # sum occurrence oer documents
        self.occurrence_doc_sum_p = self.occurrence_doc_sum_p.tolist()
        self.X_dense_binary_p = X_dense_binary_p

        #  Q distribution
        self.count_vec_q = CountVectorizer(
            ngram_range=NGRAM_RANGE,
            stop_words='english',
            lowercase=False
        )
        X_train_counts_q = self.count_vec_q.fit_transform(self.text_list_list_q)  # count occurrence per word in documents
        X_dense_q = scipy.sparse.csr_matrix.todense(X_train_counts_q)  # dense transformation
        X_dense_binary_q = scipy.sign(X_dense_q)  # binary - count 1-0 occurrence per documents
        self.occurrence_doc_sum_q = numpy.sum(X_dense_binary_q, axis=0).transpose()  # sum occurrence oer documents
        self.occurrence_doc_sum_q = self.occurrence_doc_sum_q.tolist()
        self.X_dense_binary_q = X_dense_binary_q

    def load_vocabulary_vector_aggregation(self):
        """ load vocabulary support aggregation method - KLCalc, using sklearn """
        text_file_p = open(self.description_file_p, "r")
        text_str_p = text_file_p.read().replace('\n', ' ')

        text_file_q = open(self.description_file_q, "r")
        text_str_q = text_file_q.read().replace('\n', ' ')

        Logger.info('P #of words: {}'.format(str(len(text_str_p.split()))))
        Logger.info('Q #of words: {}'.format(str(len(text_str_q.split()))))

        self.vectorizer = CountVectorizer(
            ngram_range=NGRAM_RANGE,    # (1, 2),
            stop_words='english',
            lowercase=True
        )

        self.corpus = [
            text_str_p,
            text_str_q,
        ]
        self.X = self.vectorizer.fit_transform(self.corpus)

    def calculate_kl(self):
        """ calculate KL values """
        self.X_dense = scipy.sparse.csr_matrix.todense(self.X)
        self.X_dense = self.X_dense.astype(np.float)
        self.X_dense_kl = copy.deepcopy(self.X_dense)

        # smooth zeros
        for (x, y), value in np.ndenumerate(self.X_dense):
            if value == 0:
                self.X_dense_kl[x, y] = 0.01

        # normalize vectors
        sum_of_p = sum(self.X_dense_kl[0].tolist()[0])
        self.normalize_p = [x / sum_of_p for x in self.X_dense_kl[0].tolist()[0]]
        sum_of_q = sum(self.X_dense_kl[1].tolist()[0])
        self.normalize_q = [x / sum_of_q for x in self.X_dense_kl[1].tolist()[0]]

        # calculate KL
        kl_1 = scipy.stats.entropy(self.normalize_p, self.normalize_q)
        kl_2 = scipy.stats.entropy(self.normalize_q, self.normalize_p)

        Logger.info('KL value 1: {}'.format(str(kl_1)))
        Logger.info('KL value 2: {}'.format(str(kl_2)))

    def calculate_kl_and_language_models_documents(self):
        """ calculate KL results (both), and most separate values """
        # calculate D(p||q)
        name = 'P_{}_Q_{}'.format(str(self.file_name_p), str(self.file_name_q))
        Logger.info(name)
        try:
            self._calculate_separate_words_documents(self.occurrence_doc_sum_p, self.occurrence_doc_sum_q,
                                                    self.count_vec_p.vocabulary_, self.count_vec_q.vocabulary_,
                                                    self.len_p, self.len_q, name, self.text_list_list_p,
                                                    self.text_list_list_q, self.X_dense_binary_p,
                                                    self.X_dense_binary_q, name, self.file_name_p, self.unique_buyer_p,
                                                    self.unique_buyer_q, self.df_p, self.df_q)
        except Exception, e:
            Logger.info('Exception occurred: {}'.format(e))
            Logger.info(traceback.print_exc())

        # TODO switch between p and q global values
        """
        Logger.info('')
        Logger.info('swap class variables')
        self.text_list_list_p, self.text_list_list_q = self.text_list_list_q, self.text_list_list_p
        self.df_p, self.df_q = self.df_q, self.df_p
        Logger.info('')
        """

        # calculate D(q||p)
        name = 'P_{}_Q_{}'.format(str(self.file_name_q), str(self.file_name_p))
        Logger.info(name)
        try:
            self._calculate_separate_words_documents(self.occurrence_doc_sum_q, self.occurrence_doc_sum_p,
                                                    self.count_vec_q.vocabulary_, self.count_vec_p.vocabulary_,
                                                    self.len_q, self.len_p, name, self.text_list_list_q,
                                                    self.text_list_list_p, self.X_dense_binary_q,
                                                    self.X_dense_binary_p, name, self.file_name_q, self.unique_buyer_q,
                                                    self.unique_buyer_p, self.df_q, self.df_p)
        except Exception, e:
            Logger.info('Exception occurred: {}'.format(e))
            Logger.info(traceback.print_exc())
        Logger.info('')

    def calculate_kl_and_language_models_aggregation(self):
        """ calculate KL results (both), and most separate values """
        self.calculate_kl()  # cal kl using scipy

        # cal most significant separate words - both direction
        self.calculate_separate_words_aggregation(self.X_dense[0].tolist()[0], self.X_dense[1].tolist()[0])

        self.calculate_separate_words_aggregation(self.X_dense[1].tolist()[0], self.X_dense[0].tolist()[0])

    def _calculate_separate_words_documents(self, X_p, X_q, dict_p, dict_q, len_p, len_q, name, p_text_list,
                                           q_text_list, X_dense_binary_p, X_dense_binary_q, excel_name, file_name_p,
                                            unique_buyer_p, unique_buyer_q, df_p, df_q):
        """ calculate most significance term to KL divergence """

        if CONTRIBUTE_TYPE == 'description_fraction':
            dict_ratio = self._calculate_word_contribution(dict_p, X_p, len_p, dict_q, X_q, len_q,
                                                           unique_buyer_p, unique_buyer_q, df_p, df_q,
                                                           X_dense_binary_p, X_dense_binary_q)

        elif CONTRIBUTE_TYPE == 'buyers_fraction':
            dict_ratio = self._calculate_word_contribution_buyers_fraction(
                dict_p, X_p, len_p, dict_q, X_q, len_q, unique_buyer_p, unique_buyer_q, df_p, df_q,
                X_dense_binary_p, X_dense_binary_q
            )
        else:
            raise ValueError('unknown contribute calculation type: {}'.format(CONTRIBUTE_TYPE))

        if NORMALIZE_CONTRIBUTE_FLAG:       # normalized cont.
            dict_ratio = self._normalized_contribution(dict_ratio)

        # save word contribution (KL contribution) for all words
        self._save_all_word_contribution(dict_ratio, dict_p, file_name_p)

        # save additional data to top k words + the description they appear in
        self._save_top_k_word_extend_data(dict_ratio, dict_p, X_p, len_p, dict_q, X_q, len_q, name, p_text_list, q_text_list,
                                          X_dense_binary_p, X_dense_binary_q, excel_name, unique_buyer_p, unique_buyer_q,
                                          df_p, df_q)

    @staticmethod
    def _calculate_word_contribution_old(dict_p, X_p, len_p, dict_q, X_q, len_q):
        """
        calculate all words and their contribution
        :returns: dict of word idx -> word contribution
        """

        Logger.info('calculate word contribution using fraction of description contains the token')

        dict_ratio = dict()  # contain word index and his KL contribute
        inv_p = {v: k for k, v in dict_p.iteritems()}

        # calculate word contribution
        for word_idx_p, tf_p in enumerate(X_p):
            if word_idx_p % 10000 == 0:
                Logger.info('calculate words contribution: {} / {}'.format(str(word_idx_p), str(len(X_p))))

            tf_p = np.float(tf_p[0] / len_p)            # p fraction
            word_p = inv_p[word_idx_p]                  # p word

            if word_p not in dict_q:
                tf_q = 1.0 / (len_q * SMOOTHING_FACTOR)             # word not in q distribution - using smoothing
            else:
                tf_q = np.float(X_q[dict_q[word_p]][0] / len_q)     # word appears in q

            contribute = tf_p * np.log(tf_p / tf_q)     # calculate contribution
            dict_ratio[word_idx_p] = contribute         # store value

        return dict_ratio

    @staticmethod
    def _calculate_word_contribution(dict_p, X_p, len_p, dict_q, X_q, len_q, unique_buyer_p, unique_buyer_q, df_p, df_q,
                                    X_dense_binary_p, X_dense_binary_q):
        """
        calculate all words and their contribution
        :returns: dict of word idx -> word contribution
        """

        Logger.info('calculate word contribution using fraction of description contains the token')
        Logger.info('clip number of description contains the token per buyer: {}'.format(CLIP_DESCRIPTION))

        dict_ratio = dict()  # contain word index and his KL contribute
        inv_p = {v: k for k, v in dict_p.iteritems()}

        # calculate word contribution
        for word_idx_p, tf_p in enumerate(X_p):
            if word_idx_p % 10000 == 0:
                Logger.info('calculate words contribution: {} / {}'.format(str(word_idx_p), str(len(X_p))))

            tf_p = np.float(tf_p[0] / len_p)  # p fraction
            word_p = inv_p[word_idx_p]  # p word

            # extract indexes of description contain the token
            p_token_desc_idx = np.nonzero(X_dense_binary_p[:, word_idx_p])[0]
            assert len(p_token_desc_idx) == X_p[word_idx_p][0]

            # truncate max desc per buyer
            buyer_p = df_p.iloc[p_token_desc_idx]['buyer_id'].value_counts().clip(upper=CLIP_DESCRIPTION)

            # frac of descriptions
            tf_p = np.float(sum(buyer_p.values)) / np.float(len_p)

            if word_p not in dict_q:
                tf_q = 1.0 / (len_q * SMOOTHING_FACTOR)  # word not in q distribution - using smoothing
            else:
                # tf_q = np.float(X_q[dict_q[word_p]][0] / len_q)  # word appears in q

                q_idx = dict_q[word_p]
                tf_q = X_q[dict_q[word_p]][0]
                q_token_desc_idx = np.nonzero(X_dense_binary_q[:, q_idx])[0]
                assert len(q_token_desc_idx) == tf_q

                buyer_q = df_q.iloc[q_token_desc_idx]['buyer_id'].value_counts().clip(upper=CLIP_DESCRIPTION)
                tf_q = np.float(sum(buyer_q.values)) / np.float(len_q)

            contribute = tf_p * np.log(tf_p / tf_q)  # calculate contribution
            dict_ratio[word_idx_p] = contribute  # store value

        return dict_ratio

    @staticmethod
    def _calculate_word_contribution_buyers_fraction(dict_p, X_p, len_p, dict_q, X_q, len_q,
                                                     unique_buyer_p, unique_buyer_q, df_p, df_q,
                                                     X_dense_binary_p, X_dense_binary_q):
        """
        calculate all words and their contribution using buyers fraction
        :returns: dict of word idx -> word contribution
        """

        Logger.info('calculate word contribution using fraction of buyers with description contains the token')

        dict_ratio = dict()  # contain word index and his KL contribute
        inv_p = {v: k for k, v in dict_p.iteritems()}

        # calculate word contribution
        for word_idx_p, tf_p in enumerate(X_p):
            if word_idx_p % 10000 == 0:
                Logger.info('calculate words contribution: {} / {}'.format(str(word_idx_p), str(len(X_p))))

            tf_p = tf_p[0]              # p word descriptions
            word_p = inv_p[word_idx_p]  # p word

            # extract indexes of description contain the token
            p_token_desc_idx = np.nonzero(X_dense_binary_p[:, word_idx_p])[0]
            assert len(p_token_desc_idx) == tf_p
            # get number of unique buyers
            buyer_p = len(df_p.iloc[p_token_desc_idx]['buyer_id'].value_counts().index.tolist())
            # calculate fraction of buyers
            tf_p = np.float(buyer_p) / unique_buyer_p

            if word_p not in dict_q:
                tf_q = 1.0 / (unique_buyer_q * SMOOTHING_FACTOR)  # word not in q distribution - using smoothing

            # calculate number of unique buyers
            else:
                q_idx = dict_q[word_p]
                tf_q = X_q[dict_q[word_p]][0]
                q_token_desc_idx = np.nonzero(X_dense_binary_q[:, q_idx])[0]
                assert len(q_token_desc_idx) == tf_q
                buyer_q = len(df_q.iloc[q_token_desc_idx]['buyer_id'].value_counts().index.tolist())
                tf_q = np.float(buyer_q) / unique_buyer_q

            contribute = tf_p * np.log(tf_p / tf_q)  # calculate contribution
            dict_ratio[word_idx_p] = contribute  # store value

        return dict_ratio

    @staticmethod
    def _normalized_contribution(dict_ratio):
        """ normalized KL coefficients """
        Logger.info('normalized word contribution')

        if NORMALIZE_CONTRIBUTE_TYPE == 'min_max':
            minimum = float(min(dict_ratio.values()))
            maximum = float(max(dict_ratio.values()))
            Logger.info('normalized min-max: before range - {} {}'.format(minimum, maximum))
            for idx, cont in dict_ratio.iteritems():
                dict_ratio[idx] = (cont - minimum) / float(maximum - minimum)
            Logger.info('normalized min-max: after range - {} {}'.format(
                min(dict_ratio.values()), max(dict_ratio.values())))

        elif NORMALIZE_CONTRIBUTE_TYPE == 'ratio':
            mean_contribute = sum(dict_ratio.values()) / len(dict_ratio.values())
            offset = 1.0 / mean_contribute
            Logger.info('offset value: {}'.format(str(round(offset, 3))))

            if offset < 0:
                raise ValueError('normalized factor is below zero: {}'.format(str(offset)))

            for idx, cont in dict_ratio.iteritems():
                dict_ratio[idx] = cont * offset

        else:
            raise ValueError('unknown normalized type')

        return dict_ratio

    # save all KL word contribution - word: contribution
    def _save_all_word_contribution(self, dict_idx_contribute, dict_word_index, file_name_p_distribution):
        """
        save all word contribution to KL metric after normalize them
        :param dict_idx_contribute:
        :param dict_word_index:
        :param file_name_p_distribution:
        :return:
        """

        dict_idx_contribute = sorted(dict_idx_contribute.items(), key=operator.itemgetter(1))
        dict_idx_contribute.reverse()
        Logger.info('number of words: {}'.format(str(len(dict_idx_contribute))))

        # create excel header
        book = xlwt.Workbook(encoding="utf-8")
        sheet1 = book.add_sheet('KL contribute')
        sheet1.write(0, 0, 'Word')
        sheet1.write(0, 1, 'contribute')
        sheet1.write(0, 2, 'percentile_contribute')

        row_insert = 0
        for idx, tup in enumerate(dict_idx_contribute):
            cur_word = dict_word_index.keys()[dict_word_index.values().index(tup[0])]       # get word by index
            percentile = round(1 - float(row_insert) / float(len(dict_idx_contribute)), 4)  # normalized
            try:
                sheet1.write(row_insert + 1, 0, str(cur_word.encode('utf8')))
                sheet1.write(row_insert + 1, 1, str(round(tup[1], 4)))
                sheet1.write(row_insert + 1, 2, str(percentile))
                row_insert += 1
            except Exception, e:
                Logger.info('Failed: {}'.format(str(e)))
                Logger.info('Failed: {} {}'.format(str(idx), str(cur_word.encode('utf8'))))

        # interesting parameters:
        # p str(self.len_p)
        # q str(self.len_q)
        # num token in file str(row_insert)

        dir_name = '{}'.format(self.dir_all_words_contribute)

        if not os.path.exists(dir_name):
            os.makedirs(dir_name)

        excel_file_name = '{}{}_{}.xls'.format(
            dir_name, str(self.pt), str(file_name_p_distribution)
        )

        Logger.info('save all words contribute in file: {}'.format(str(excel_file_name)))
        book.save(excel_file_name)

    # save top k words relevant  data
    def _save_top_k_word_extend_data(self, dict_ratio, dict_p, X_p, len_p, dict_q, X_q, len_q, name, p_text_list, q_text_list,
                                     X_dense_binary_p, X_dense_binary_q, excel_name, unique_buyer_p, unique_buyer_q, df_p, df_q):
        # save all term contribute in a file
        dict_max_ratio = sorted(dict_ratio.items(), key=operator.itemgetter(1))
        dict_max_ratio.reverse()

        # find top k tokens
        counter = 0

        # create excel header
        book = xlwt.Workbook(encoding="utf-8")

        xlwt.add_palette_colour("red_flag", 0x21)
        book.set_colour_RGB(0x21, 250,128,114)
        style = xlwt.easyxf('pattern: pattern solid, fore_colour red_flag')
        # name = name[:5]

        if len(name) > 31:
            sheet_name = name[-31:]
            sheet1 = book.add_sheet(sheet_name)
        else:
            sheet1 = book.add_sheet(name)
        sheet1.write(0, 0, 'Word')
        sheet1.write(0, 1, 'TF_P')
        sheet1.write(0, 2, 'TF_Q')
        sheet1.write(0, 3, 'Contribute')
        sheet1.write(0, 4, 'Fraction_P')
        sheet1.write(0, 5, 'Fraction_Q')
        sheet1.write(0, 6, 'Unique buyers P')
        sheet1.write(0, 7, 'Unique buyers Q')
        sheet1.write(0, 8, 'Fraction unique buyers P')
        sheet1.write(0, 9, 'Fraction unique buyers Q')
        sheet1.write(0, 10, 'Percentile contribution')

        for idx, tup in enumerate(dict_max_ratio):

            counter += 1

            cur_word = dict_p.keys()[dict_p.values().index(tup[0])]

            tf_p = X_p[tup[0]][0]

            # word does not appears in q distribution
            if cur_word not in dict_q:
                tf_q = 0
                q_idx = -1

            # word appears in q
            else:
                tf_q = X_q[dict_q[cur_word]][0]
                q_idx = dict_q[cur_word]

            frac_p = np.float(tf_p) / len_p
            frac_q = np.float(tf_q) / len_q

            # extract indexes of description contain the token
            p_token_desc_idx = np.nonzero(X_dense_binary_p[:, tup[0]])[0]
            assert len(p_token_desc_idx) == tf_p

            # get number of unique buyers
            buyer_p = len(df_p.iloc[p_token_desc_idx]['buyer_id'].value_counts().index.tolist())

            # calculate fraction of buyers
            frac_buyer_p = np.float(buyer_p) / unique_buyer_p

            if tf_q != 0:
                q_token_desc_idx = np.nonzero(X_dense_binary_q[:, q_idx])[0]
                assert len(q_token_desc_idx) == tf_q
                buyer_q = len(df_q.iloc[q_token_desc_idx]['buyer_id'].value_counts().index.tolist())
                frac_buyer_q = np.float(buyer_q) / unique_buyer_q
            else:
                buyer_q, frac_buyer_q = 0, 0

            percentile = round(1 - float(idx) / float(len(dict_max_ratio)), 4)  # normalized

            sheet1.write(idx + 1, 0, str(cur_word)) if frac_buyer_p > 0.05 else sheet1.write(idx + 1, 0, str(cur_word), style)
            sheet1.write(idx + 1, 1, str(tf_p))
            sheet1.write(idx + 1, 2, str(tf_q))
            sheet1.write(idx + 1, 3, str(round(tup[1], 2)))
            sheet1.write(idx + 1, 4, str(round(frac_p, 3)))
            sheet1.write(idx + 1, 5, str(round(frac_q, 3)))
            sheet1.write(idx + 1, 6, str(buyer_p))
            sheet1.write(idx + 1, 7, str(buyer_q))
            sheet1.write(idx + 1, 8, str(round(frac_buyer_p, 3)))
            sheet1.write(idx + 1, 9, str(round(frac_buyer_q, 3)))
            sheet1.write(idx + 1, 10, str(percentile))

            if FIND_WORD_DESCRIPTION_FLAG and counter < TOP_K_WORDS:
                # find occurrences of current terms and save descriptions in both distribution in excel file
                self.find_occurrences_current_terms(str(cur_word), tup, q_idx, tf_p, tf_q, p_text_list, q_text_list,
                                                    X_dense_binary_p, X_dense_binary_q, excel_name, counter, df_p, df_q)

        # save top k token with counting

        dir_name = '{}'.format(self.dir_kl_detailed_word)
        if not os.path.exists(dir_name):
            os.makedirs(dir_name)

        excel_file_name = '{}{}_K_{}_Smooth_{}_{}_{}_top_{}_{}.xls'.format(
            dir_name, self.pt, str(TOP_K_WORDS), str(SMOOTHING_FACTOR),
            str(VOCABULARY_METHOD), str(name), str(TOP_K_WORDS),  str(self.cur_time)
        )

        book.save(excel_file_name)
        Logger.info('save kl detailed data in file: {}'.format(str(excel_file_name)))
        Logger.info('')

    def calculate_separate_words_aggregation(self, X_p, X_q):
        """ calculate most significance term to KL divergence """
        self.X_dense_p = X_p    # self.X_dense[0].tolist()[0]
        self.X_dense_q = X_q    # self.X_dense[1].tolist()[0]

        sum_of_p = np.float(sum(self.X_dense_p))
        sum_of_q = np.float(sum(self.X_dense_q))

        dict_ratio = dict()     # contain word index and his KL contribute

        for word_idx, tf_p in enumerate(self.X_dense_p):

            if tf_p > 0:
                tf_p = np.float(tf_p)/sum_of_p
                tf_q = np.float(self.X_dense_q[word_idx])/sum_of_q

                # smoothing
                if tf_q == 0:
                    tf_q = 1.0 / (sum_of_q * SMOOTHING_FACTOR)

                contribute = tf_p * np.log(tf_p/tf_q)
                dict_ratio[word_idx] = contribute

        # find top k
        dict_max_ratio = sorted(dict_ratio.items(), key=operator.itemgetter(1))
        dict_max_ratio.reverse()
        counter = 0

        for idx, tup in enumerate(dict_max_ratio):

            counter += 1
            if counter > TOP_K_WORDS:
                break

            tf_p = self.X_dense_p[tup[0]]
            tf_q = self.X_dense_q[tup[0]]
            frac_p = np.float(tf_p) / sum_of_p
            frac_q = np.float(tf_q) / sum_of_q
            cur_ratio = self.vectorizer.vocabulary_.keys()[self.vectorizer.vocabulary_.values().index(tup[0])]

            Logger.debug('{}, tf p: {}, tf q: {}, cont: {}, ratio_tf p: {}, ratio_tf q: {}'.format(
                str(cur_ratio), str(tf_p), str(tf_q), str(round(tup[1], 2)),
                str(round(frac_p, 3)), str(round(frac_q, 3))
            ))
        Logger.info('')

    # calculate term with largest ratio and present them
    def cal_ratio_tf_q_p(self, dict_max_ratio):
        counter = 0
        for idx, tup in enumerate(dict_max_ratio):
            counter += 1
            if counter > TOP_K_WORDS:
                break
            tf_p = self.X_dense.tolist()[0][tup[0]]
            tf_q = self.X_dense.tolist()[1][tup[0]]
            cur_ratio = self.vectorizer.vocabulary_.keys()[self.vectorizer.vocabulary_.values().index(tup[0])]
            Logger.denbug('{}, tf p: {}, tf q: {}'.format(str(cur_ratio), str(tf_p), str(tf_q)))

    def find_occurrences_current_terms(self, cur_word, tup, q_idx, tf_p, tf_q, p_text_list, q_text_list,
                                       X_dense_binary_p, X_dense_binary_q, excel_name, counter_top_idx, df_p, df_q):
        """
        high time complexity
        :param cur_word: input term - we seek descriptions contain the terms
        :param tf_p: number of description in p contain input term
        :param tf_q: number of description in p contain input term
        :return:
        examples of descriptions which contain term input
        """

        assert len(p_text_list) == X_dense_binary_p.shape[0]
        assert len(q_text_list) == X_dense_binary_q.shape[0]

        book = xlwt.Workbook(encoding="utf-8")

        # P distribution
        sheet1 = book.add_sheet('P_{}'.format(str(tf_p)), cell_overwrite_ok=True)
        sheet1.write(0, 0, 'Item index')
        sheet1.write(0, 1, 'Item description')
        sheet1.write(0, 2, 'buyer_id')

        row_p_i = 0
        cur_word_doc_counter = 0
        cnt = 0
        for row in X_dense_binary_p:                # iterate over description (row=desc)
            if cnt > FIND_WORD_DESCRIPTION_K:
                break
            cur_row_list = np.array(row)[0].tolist()    # binary list - 1 if word in descriptions
            if cur_row_list[tup[0]] > 0:                # iterate over all description, extract where '1' on the token
                cur_word_doc_counter += 1               # find desc with the token
                cnt += 1
                sheet1.write(cur_word_doc_counter, 0, row_p_i)

                assert str(cur_word) in p_text_list.iloc[row_p_i]
                assert df_p.iloc[row_p_i][KL_TARGET_COLUMN] == p_text_list.iloc[row_p_i]

                sheet1.write(cur_word_doc_counter, 1, p_text_list.iloc[row_p_i])
                sheet1.write(cur_word_doc_counter, 2, str(df_p.iloc[row_p_i]['buyer_id']))
            row_p_i += 1

        # assert cur_word_doc_counter == tf_p

        if q_idx >= 0:
            # Q distribution
            sheet2 = book.add_sheet('Q_{}'.format(str(tf_q)))
            sheet2.write(0, 0, 'Item index')
            sheet2.write(0, 1, 'Item description')
            sheet2.write(0, 2, 'buyer_id')
            row_q_i = 0
            cnt = 0
            cur_word_doc_counter = 0
            for row in X_dense_binary_q:
                if cnt > FIND_WORD_DESCRIPTION_K:
                    break
                cur_row_list = np.array(row)[0].tolist()  # binary list - 1 if word in descriptions
                if cur_row_list[q_idx] > 0:
                    cur_word_doc_counter += 1
                    sheet2.write(cur_word_doc_counter, 0, row_q_i)

                    assert str(cur_word) in q_text_list.iloc[row_q_i]
                    assert df_q.iloc[row_q_i][KL_TARGET_COLUMN] == q_text_list.iloc[row_q_i]

                    sheet2.write(cur_word_doc_counter, 1, q_text_list.iloc[row_q_i])
                    sheet2.write(cur_word_doc_counter, 2, str(df_q.iloc[row_q_i]['buyer_id']))
                    cnt += 1
                row_q_i += 1

            # assert cur_word_doc_counter == tf_q

        # directory with file to each token
        cur_dir = '{}{}_{}_{}/'.format(
            self.dir_excel_token_appearance, str(self.pt), str(excel_name), str(self.cur_time))

        if not os.path.exists(cur_dir):
            os.makedirs(cur_dir)

        excel_file_name = '{}{}_{}_P_{}_Q_{}_{}.xls'.format(
            cur_dir, str(counter_top_idx), str(cur_word), str(tf_p), str(tf_q), str(self.cur_time)
        )

        book.save(excel_file_name)
        Logger.debug('save descriptions for top-k token in file: ' + str(excel_file_name))

    def _plot_histogram(self, a, user_amount, desc_num, additional_str, pt, bin=50):
        """ generic file to plot histogram and save plot """

        plt_dir = '../results/pre-processing/calculate_kl/{}/'.format(self.cur_time)
        if not os.path.exists(plt_dir):
            os.makedirs(plt_dir)

        plt.style.use('seaborn-deep')
        plt.hist(a, bins=bin)
        plt.title('Description per user - {}, {}'.format(additional_str, pt))

        plot_path = '{}{}_user_{}_desc_{}_group_{}.png'.format(
            plt_dir,
            pt,                 # personality trait
            str(user_amount),
            str(desc_num),
            additional_str
        )

        plt.savefig(plot_path)
        plt.close()

        Logger.info('save histogram plot: {}'.format(str(plot_path)))

    def _save_config_values(self):
        """ save configuration into file"""
        with open('{}config.txt'.format(self.dir_config), 'w') as file:
            file.write(json.dumps(KL_CONFIGURATION_DICT))


def main(merge_df_path, log_file_name):

    # init class
    create_vocabularies_obj = CalculateKL(
        merge_df_path=merge_df_path,
        log_file_name=log_file_name,
        level='info'
    )

    create_vocabularies_obj.check_input()                       # check if arguments are valid
    create_vocabularies_obj.run_kl()                            # contain all inner functions


if __name__ == '__main__':

    # extraversion
    # description_file_p = '../results/vocabulary/extraversion/documents_high_extraversion_534_2018-06-11 19:29:07.txt'
    # description_file_q = '../results/vocabulary/extraversion/documents_low_extraversion_939_2018-06-11 19:29:07.txt'
    # description_file_p = '../results/vocabulary/extraversion/documents_high_extraversion_500_2018-06-13 12:00:21.txt'
    # description_file_q = '../results/vocabulary/extraversion/documents_low_extraversion_885_2018-06-13 12:00:21.txt'
    # description_file_p = '../results/vocabulary/extraversion/documents_high_extraversion_410_2018-06-17 09:04:28.txt'
    # description_file_q = '../results/vocabulary/extraversion/documents_low_extraversion_876_2018-06-17 09:04:28.txt'
    # description_file_p = '../results/vocabulary/extraversion/documents_high_extraversion_457_2018-06-17 09:13:38.txt'
    # description_file_q = '../results/vocabulary/extraversion/documents_low_extraversion_1106_2018-06-17 09:13:38.txt'
    # description_file_p = '../results/vocabulary/extraversion/documents_high_extraversion_962_2018-06-17 15:48:07.txt'
    # description_file_q = '../results/vocabulary/extraversion/documents_low_extraversion_2175_2018-06-17 15:48:07.txt'

    # openness
    # description_file_p = '../results/vocabulary/openness/documents_high_openness_2018-06-10 07:28:45.txt'
    # description_file_q = '../results/vocabulary/openness/documents_low_openness_2018-06-10 07:28:45.txt'
    # trait = 'openness'

    # agreeableness
    # description_file_p = '../results/vocabulary/agreeableness/documents_high_agreeableness_2018-06-10 07:57:48.txt'
    # description_file_q = '../results/vocabulary/agreeableness/documents_low_agreeableness_2018-06-10 07:57:48.txt'
    # trait = 'agreeableness'

    # conscientiousness
    # description_file_p = '../results/vocabulary/conscientiousness/documents_high_conscientiousness_2018-06-10 07:53:27.txt'
    # description_file_q = '../results/vocabulary/conscientiousness/documents_low_conscientiousness_2018-06-10 07:53:27.txt'
    # trait = 'conscientiousness'

    # neuroticism
    # description_file_p = '../results/vocabulary/neuroticism/documents_high_neuroticism_2018-06-10 07:38:36.txt'
    # description_file_q = '../results/vocabulary/neuroticism/documents_low_neuroticism_2018-06-10 07:38:36.txt'
    # description_file_p = '../results/vocabulary/neuroticism/documents_high_neuroticism_705_2018-06-17 09:19:06.txt'
    # description_file_q = '../results/vocabulary/neuroticism/documents_low_neuroticism_858_2018-06-17 09:19:06.txt'
    # trait = 'neuroticism'

    # merge_df_path = '../results/data/vocabularies/5505_2018-08-04 21:17:50.csv'
    file_prefix = 'run_kl'
    log_file_name = '../log/{}_{}.log'.format(file_prefix, strftime("%Y-%m-%d %H:%M:%S", gmtime()))
    # merge_df_path = '../results/data/vocabularies/5457_2018-08-07 15:06:22.csv'
    merge_df_path = '../results/data/vocabularies/3954_2018-08-12 07:50:59.csv'
    main(merge_df_path, log_file_name)
